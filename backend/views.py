import logging

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import viewsets
from rest_framework.decorators import action

from backend.models import ShipmentLog
from backend.pagination import ShipmentlogPagination
from backend.serializers import ShipmentLogSerializer

# Create your views here.

logger = logging.getLogger(__name__)


class ShipmentLogViewSet(viewsets.ModelViewSet):
    """Контроллер для страницы SPA-приложения"""

    queryset = ShipmentLog.objects.all()
    serializer_class = ShipmentLogSerializer
    pagination_class = ShipmentlogPagination

    def get_queryset(self):
        queryset = super().get_queryset()

        # Считывание параметров фильтрации из URL-запроса
        column = self.request.query_params.get("column")  # выбор колонки (название, количество, расстояние)
        condition = self.request.query_params.get("condition")  # условие (равно, содержит, больше, меньше)
        value = self.request.query_params.get("value")  # значение для фильтрации

        if column and condition and value:
            lookup = None
            if condition == "equals":
                lookup = f"{column}__iexact" if column == "name" else f"{column}"
            elif condition == "contains" and column == "name":
                lookup = "name__icontains"
            elif condition == "greater":
                lookup = f"{column}__gt"
            elif condition == "less":
                lookup = f"{column}__lt"

            if lookup:
                try:
                    queryset = queryset.filter(**{lookup: value})
                    logger.info(f"Применен фильтр: {lookup}={value}. Найдено записей: {queryset.count()}")
                except (TypeError, ValueError) as e:
                    logger.warning(
                        f"Ошибка валидации типа данных при фильтрации. "
                        f"Колонка: {column}, Условие: {condition}, Значение: {value}. Ошибка: {e}"
                    )
                    pass

        ordering = self.request.query_params.get("ordering")
        if ordering and "date" not in ordering:
            queryset = queryset.order_by(ordering)

        return queryset

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        """Экспорт всех отфильтрованных записей в файл Excel"""
        # 1. Получаем отфильтрованный QuerySet
        queryset = self.get_queryset()

        # 2. Создаем книгу Excel в памяти
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Рейсы LogiTrack"

        # Стили оформления
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E4ED8", end_color="1E4ED8", fill_type="solid")  # Синий градиент
        data_font = Font(name="Segoe UI", size=10)
        center_align = Alignment(horizontal="center", vertical="center")

        # 3. Записываем заголовки
        headers = ["Дата рейса", "Маршрут", "Количество груза (ед.)", "Расстояние (км)"]
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # 4. Записываем данные из БД
        for shipment in queryset:
            # Форматируем дату
            date_str = shipment.date.strftime("%d.%m.%Y %H:%M") if shipment.date else ""

            row_data = [date_str, shipment.name, shipment.quantity, shipment.distance]

            # Записываем строку в таблицу Excel
            ws.append(row_data)

            # Применяем созданный data_font к каждой ячейке новой строки
            current_row = ws.max_row
            for col_num in range(1, 5):
                ws.cell(row=current_row, column=col_num).font = data_font

        # Выравниваем ширину колонок под текст автоматически
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # 5. Формируем HTTP-ответ с файлом
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename='LogiTrack_Full_Report.xlsx'"

        # Записываем книгу в тело ответа
        wb.save(response)
        return response
